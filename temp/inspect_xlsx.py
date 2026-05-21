import openpyxl
p = r'C:\Users\asus\OneDrive\Desktop\seğmen su\excel data\1NALKAYA-2026-17 OCAK DAMACANA DAMACANA.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
for ws in wb.worksheets:
    print('SHEET', ws.title, ws.max_row, ws.max_column)
    for r in ws.iter_rows(min_row=1, max_row=min(12, ws.max_row), values_only=True):
        print([str(x) if x is not None else '' for x in r[:20]])
    print('---')
