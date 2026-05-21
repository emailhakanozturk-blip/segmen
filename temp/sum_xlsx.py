import openpyxl
p=r'C:\Users\asus\OneDrive\Desktop\seğmen su\excel data\1NALKAYA-2026-17 OCAK DAMACANA DAMACANA.xlsx'
wb=openpyxl.load_workbook(p,data_only=True)
ws=wb['OCAK']
rows=[]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[1] and str(r[1]).startswith('SEI'):
        rows.append(r)
print(len(rows), sum(float(r[12] or 0) for r in rows), sum(float(r[13] or 0) for r in rows), sum(float(r[14] or 0) for r in rows), sum(float(r[15] or 0) for r in rows))
print('totalrow', [ws.cell(51,c).value for c in range(13,17)])
