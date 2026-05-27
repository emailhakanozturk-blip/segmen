import json
import sys
from datetime import datetime

import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


def cell_text(value):
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")

    return str(value).strip()


def sheet_score(sheet):
    score = 0

    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and cell.strip().upper().startswith("SEI"):
                score += 1

    return score


def tariff_rows(workbook):
    if "Fiyatlar" not in workbook.sheetnames or "Motorin Fiyat Tablosu" not in workbook.sheetnames:
        return []

    motorin_sheet = workbook["Motorin Fiyat Tablosu"]
    base_date = ""
    base_motorin = ""

    for row in motorin_sheet.iter_rows(values_only=True):
        values = list(row)
        if len(values) >= 3 and isinstance(values[1], datetime) and isinstance(values[2], (int, float)):
            base_date = values[1].strftime("%d.%m.%Y")
            base_motorin = str(values[2])
            break

    if base_motorin == "":
        return []

    rows = []
    fiyat_sheet = workbook["Fiyatlar"]

    for row in fiyat_sheet.iter_rows(values_only=True):
        values = list(row)
        if len(values) < 4:
            continue

        if isinstance(values[1], datetime) and len(values) >= 7:
            cikis = cell_text(values[2])
            varis = cell_text(values[3])
            birim = values[4]
            motorin = values[6] if isinstance(values[6], (int, float)) else base_motorin
        else:
            cikis = cell_text(values[1])
            varis = cell_text(values[2])
            birim = values[3]
            motorin = base_motorin

        if cikis == "" or varis == "" or not isinstance(birim, (int, float)):
            continue

        rows.append(["__TARIFE__", cikis, varis, str(birim), str(motorin), base_date])

    return rows


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Dosya yolu eksik.")

    with open(sys.argv[1], "rb") as source:
        workbook = openpyxl.load_workbook(source, data_only=True)
    sheet = max(workbook.worksheets, key=sheet_score)

    rows = tariff_rows(workbook)
    for row in sheet.iter_rows(values_only=True):
        values = [cell_text(value) for value in row]

        while values and values[-1] == "":
            values.pop()

        if any(value != "" for value in values):
            rows.append(values)

    print(json.dumps(rows, ensure_ascii=False))


if __name__ == "__main__":
    main()
